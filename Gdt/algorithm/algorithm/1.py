# -*- coding: utf-8 -*-
"""
Created on Wed Oct 16 09:45:08 2019

@author: hsc
"""
import matplotlib.pyplot as plt
from openpyxl import load_workbook

workbook = load_workbook(u'C:/Users/hsc/Desktop/liver.xlsx')    #找到需要xlsx文件的位置
booksheet = workbook.active                 #获取当前活跃的sheet,默认是第一个sheet

#如果想获取别的sheet页采取下面这种方式，先获取所有sheet页名，在通过指定那一页。
# sheets = workbook.get_sheet_names()  # 从名称获取sheet
# booksheet = workbook.get_sheet_by_name(sheets[0])

#获取sheet页的行数据
rows = booksheet.rows
#获取sheet页的列数据
columns = booksheet.columns


data= []
i=2
columns= 2
for i in rows:
    while columns<= 11 :
    #将表中第二行行的1-11列数据写入data数组中
        m=data.append(booksheet.cell(row=2, column=columns).value)
        colunmns=columns+1
    


fig  = plt.figure()

ax = fig.add_subplot(1,1,1)

ax.plot([4,13],m)
    
ax.set_title("Correlation")
ax.set_xlabel("x-week")
ax.set_ylabel("expression")

print('fit')